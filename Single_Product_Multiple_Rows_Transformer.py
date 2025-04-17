import dearpygui.dearpygui as dpg
import openpyxl
import re
from openpyxl.utils import get_column_letter, coordinate_to_tuple
import os
import time
import gspread
from Google_Sheets_and_Drive_Auth import auth_gspread

def display_single_product_multiple_rows_transformer(sender,app_data,user_data):

    filepath=user_data

    with dpg.window(label="Transform Spreadsheet With Multiple Rows for Each Product",width=600,height=250) as transformer_window:
        
        dpg.add_progress_bar(id=f'{transformer_window}_progress', overlay="Loading formatting dicts...", show=True, width=200, parent=transformer_window)
        format_dicts = parse()
        dpg.configure_item(f'{transformer_window}_progress', show=True, default_value=1, overlay=f"{len(format_dicts)} formatting dict(s) loaded.")
        
        input_group = dpg.add_group(horizontal=True)
        dpg.add_button(tag=f'{transformer_window}_input_fileselect', label="Choose Input File", callback=file_select, user_data=transformer_window, parent=input_group, width=150)
        dpg.add_input_text(tag=f'{transformer_window}_input_filename', label="Input File", parent=input_group, width=350)
        dpg.add_input_text(tag=f'{transformer_window}_output_filename', label="Output File", width=500)
        dpg.add_button(label="Convert to One Row Per Product", callback=begin_convert, user_data=format_dicts, width=400)
        

def file_select(sender, app_data, user_data):
    with dpg.file_dialog(modal=True, default_path=os.path.join(os.path.expanduser('~'), 'downloads'), callback=update_file_select, height=400, width=400) as dialog:
        dpg.add_file_extension(".xlsx", color=(255, 255, 0, 255), custom_text="XLSX")
        dpg.set_item_user_data(dialog, user_data)

def update_file_select(sender, app_data, user_data):
    sender_parent = user_data
    dpg.set_value(f'{sender_parent}_input_filename', app_data['file_path_name'])
    if not dpg.get_value(f'{sender_parent}_output_filename'):
        output_filename = os.path.join(dpg.get_value('base_input_path'), app_data['file_name'])
        if app_data['file_path_name'] != output_filename:
            dpg.set_value(f'{sender_parent}_output_filename', output_filename)

def begin_convert(sender, app_data, user_data):

    transformer_window = dpg.get_item_parent(sender)
    format_dicts = user_data
    input_filename = dpg.get_value(f'{transformer_window}_input_filename')
    output_filename = dpg.get_value(f'{transformer_window}_output_filename')
    progress_bar_id = f'{transformer_window}_progress'

    dpg.configure_item(progress_bar_id, show=True, default_value=0, overlay=f"0%")

    wb = openpyxl.load_workbook(input_filename, data_only=True)
    ws = wb.active

    matched_format_dict = None
    
    for name, format_dict in format_dicts.items():
        if format_dict.match(ws):
            dpg.configure_item(progress_bar_id, show=True, default_value=0.1, overlay=f"10%")
            matched_format_dict = format_dict
            break

    if matched_format_dict is None:
        dpg.configure_item(progress_bar_id, show=True, default_value=0, overlay="No formatting dict found!")
        return

    output_wb = openpyxl.Workbook()
    output = output_wb.active

    dpg.configure_item(progress_bar_id, show=True, default_value=0.2, overlay="20%")

    for column_number, mapping in enumerate(format_dict.values):
        output.cell(row=1, column=column_number + 1).value = mapping.output

    rows_to_skip = format_dict.header_line_count
    rows = []
    output_row_number = 2

    for row in ws.iter_rows():
        if rows_to_skip != 0:
            rows_to_skip -= 1
            continue
        rows.append(row)
        if len(rows) == format_dict.lines_per_product:
            r = ""
            any = False
            for column_number, mapping in enumerate(format_dict.values):
                for coord in mapping.coords:
                    value_row, value_column = coordinate_to_tuple(coord)
                    value = normalize(rows[value_row - 1][value_column - 1].value) or ""
                    if mapping.regex:
                        match = mapping.regex.search(value)
                        if not match:
                            value = ""
                        else:
                            value = match.group()
                    if value:
                        r = r + " " + value
                        output.cell(row=output_row_number, column=column_number + 1).value = value
                        any = True
                        break
            rows = []
            if any:
                output_row_number += 1
    dpg.configure_item(progress_bar_id,show=True,default_value=0.9,overlay="90%")    
    output_wb.save(output_filename)
    dpg.configure_item(progress_bar_id,show=True,default_value=1,overlay="Done")

def parse():

    gc = auth_gspread()
    sh = gc.open("JFGC Single Product Multiple Rows Transformer")

    format_dicts = {}

    for wk in sh.worksheets():

        list_of_lists = wk.get_all_values()

        locations = None
        headings = []
        mappings = []
        header_line_count = None
        lines_per_product = None

        for row_number, row in enumerate(list_of_lists):
            print(row)
            if locations is None:
                locations = Locations(row)
                print(f"{wk.title}: {locations}")
            else:

                type = normalize(row[locations.type])
                if type is None:
                    raise ValueError(f"{wk.title}: No type found on row {row_number + 1}")

                coords = normalize(row[locations.coords])
                if coords is None:
                    raise ValueError(f"{wk.title}: No location found on row {row_number + 1}")
                coords = re.findall("[A-Z]+\\d+", coords)

                regex = normalize(row[locations.regex])
                if regex is not None:
                    regex = re.compile(regex)

                if row_number == 1:
                    header_line_count_string = normalize(row[locations.header_line_count])
                    if header_line_count_string is None:
                        raise ValueError(f"{wk.title}: Header Line Count not found!")
                    header_line_count = int(header_line_count_string)

                    lines_per_product_string = normalize(row[locations.lines_per_product])
                    if lines_per_product_string is None:
                          raise ValueError(f"{wk.title}: Lines Per Product not found!")
                    lines_per_product = int(lines_per_product_string)

                if type == "Key":
                    if regex is None:
                        raise ValueError(f"{wk.title}: No regex found on row {row_number + 1}")
                    headings.append(Heading(
                        coords,
                        regex
                    ))
                elif type == "Value":
                    output = normalize(row[locations.output])
                    if output is None:
                        raise ValueError(f"{wk.title}: No output found on row {row_number + 1}")
                    mappings.append(Mapping(
                        coords,
                        regex,
                        output
                    ))
                else:
                    raise ValueError(f"Invalid type {type}")
        
        if len(mappings) == 0:
            raise ValueError(f"{wk.title}: No mappings found!")

        format_dicts[wk.title] = FormatDict(headings, mappings, header_line_count, lines_per_product)
    return format_dicts

def normalize(value):
    if value is None:
        return None
    else:
        stripped = str(value).strip()
        return stripped or None

class Locations:
    type: int
    coords: int
    regex: int
    output: int
    header_line_count: int
    lines_per_product: int

    def __init__(self, row):

        for column_number, cell in enumerate(row):
            name = normalize(cell)
            if name is not None:
                if name == "Type":
                    self.type = column_number
                elif name == "Location":
                    self.coords = column_number
                elif name == "Regex":
                    self.regex = column_number
                elif name == "Output":
                    self.output = column_number
                elif name == "Header Line Count":
                    self.header_line_count = column_number
                elif name == "Lines Per Product":
                    self.lines_per_product = column_number
                else:
                    raise ValueError(f"{ws.title}: Unrecognized heading {val}")

        if self.type is None:
            raise ValueError("Type column not found!")
        if self.coords is None:
            raise ValueError("Location column not found!")
        if self.regex is None:
            raise ValueError("Regex column not found!")
        if self.output is None:
            raise ValueError("Output column not found!")
        if self.header_line_count is None:
            raise ValueError("Header line Count column not found!")
        if self.lines_per_product is None:
            raise ValueError("Lines Per Product column not found!")

    def __repr__(self):
        return f"{self.type} {self.coords} {self.regex} {self.output} {self.header_line_count} {self.lines_per_product}"

class Mapping:

    coords: list[str]
    regex: re.Pattern
    output: str

    def __init__(self, coords, regex, output):
        self.coords = coords
        self.regex = regex
        self.output = output

    def __repr__(self):
        regex = f" with {self.regex}" if self.regex else ""
        return f"{self.coords} @ {self.coords}{kr} -> {self.output}"

class Heading:

    coords: list[str]
    regex: re.Pattern

    def __init__(self, coords, regex):
        self.coords = coords
        self.regex = regex

    def __repr__(self):
        regex = f" with {self.regex}" if self.regex else ""
        return f"{self.coords} @ {self.coords}{kr}"

class FormatDict:

    headings: list[Heading]
    values: list[Mapping]
    header_line_count: int
    lines_per_product: int

    def __init__(self, headings, values, header_line_count, lines_per_product):
        self.headings = headings
        self.values = values
        self.header_line_count = header_line_count
        self.lines_per_product = lines_per_product

    def __repr__(self):
        return f"{self.header_line_count} {self.lines_per_product} {self.headings} {self.values} "

    @property
    def full_name(self):
        return self.name if self.is_rubric else f'{self.name}_{self.subname}'

    def match(self, worksheet):
        for heading in self.headings:
            match_found = False
            for coord in heading.coords:
                key_row, key_column = coordinate_to_tuple(coord)
                check = normalize(worksheet[key_row][key_column - 1].value) or ""
                print(f"check: {check}, regex: {heading.regex}")
                if heading.regex.search(check):
                    match_found = True
                    break
            if not match_found:
                return False
        return True