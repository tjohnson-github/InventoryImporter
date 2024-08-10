
import pickle
import openpyxl
import csv
import sys,os
import pandas as pd
import tabula



def mkdirWrapper(path):
   
    if not os.path.exists(path):
        os.mkdir(path) 
    
#/////////////////////////////////////////////////////
#           EXCEL / CSV <--> LIST CONVERSIONS
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
    
# Replace the following with raised exceptions!

def csv_to_list(csv_file,delimiter = ','):
	#-------------------------------------------
    if os.path.exists(csv_file)==False:     
        raise Exception(f"File Path does not exist for item: {str(csv_file)}")
    elif csv_file.endswith(".csv")==False:
        raise Exception(f"{str(csv_file)} is not correct file type for this operation; .csv only please.")
	#-------------------------------------------
    readArray=[]
    with open(csv_file,newline='') as csvfile:
        contents=csv.reader(csvfile, delimiter=delimiter, quotechar='"')
        for row in contents:
            temp_array=[]
            #----------
            for item in row:
                #============ TRIM WHITESPACES!
                temp_value = str(item)
                while temp_value.endswith(' '):
                    temp_value = temp_value[:-1]
                temp_array.append(item)
                #============ TRIM WHITESPACES!
            #----------
            if (temp_array != [None for i in range(0,len(temp_array))]) and (temp_array != ['' for i in range(0,len(temp_array))]):
                readArray.append(temp_array)
	#-------------------------------------------
    return readArray


def excel_to_list(excel_file):
	#-------------------------------------------
    if os.path.exists(excel_file)==False:   
        raise Exception(f"File Path does not exist for item: {str(excel_file)}")
    elif not excel_file.endswith('.xlsx'):
        raise Exception(f"{str(excel_file)} is not correct file type for this operation; .xlsx only please.")
	#-------------------------------------------
    wb = openpyxl.load_workbook(excel_file,data_only=True)
    ws = wb.active
	#-------------------------------------------
    full_list=[]
	#-------------------------------------------
    for i,row in enumerate(ws.iter_rows()):
        #----------------
        temp_list=[]
        #----------------
        for cell in row:
            #----------------
            if cell.value==' ':     
                temp_list.append('')
            else:                   
                #============ TRIM WHITESPACES!
                temp_value = str(cell.value)
                while temp_value.endswith(' '):
                    temp_value = temp_value[:-1]
                temp_list.append(temp_value)
                #============ TRIM WHITESPACES!
        #----------------
        if i==0:
            noneCount = 0
            for ii,item in enumerate(temp_list):
                if item==None or item=="None":
                    noneCount+=1
                else: 
                    noneCount=0

                if ii==len(temp_list)-1 and noneCount>0:
                    temp_list = temp_list[:ii-noneCount+1]
        else:
            temp_list = temp_list[:len(temp_list)-noneCount]

        #----------------
        if (temp_list != [None for i in range(0,len(temp_list))]) and (temp_list != ['' for i in range(0,len(temp_list))]):
            full_list.append(temp_list)
	#-------------------------------------------
    return full_list


def list_to_excel(temp_list,filename_to_save):
    import openpyxl
    try:
        #------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        #------------------------
        for row in temp_list:
            print(row)
            ws.append(row)
        #------------------------
        wb.save(filename_to_save)
        return True
    except Exception as e:
        print (f"{filename_to_save} not saved correctly:\t{e}")
        return False

def list_to_csv(temp_list, filename_to_save, delimiter = ','):
    try:
        #------------------------
        with open(filename_to_save, 'w', newline='') as csvfile:
            writer_object = csv.writer(csvfile, delimiter = delimiter, quotechar='"', quoting = csv.QUOTE_MINIMAL)
            #------------------------
            for row in temp_list:
                    try:
                        writer_object.writerow(row)
                    except Exception as e:
                        print(f"PROBLEMATIC ROW:\n\t{row}")

                        for i,x in enumerate(row):
                            if type(x)==str:
                                if 'D\x90' in x: 
                                    row[i] = x.replace('D\x90','DE')

                                    try:
                                        writer_object.writerow(row)
                                        break
                                    except Exception as e:
                                        print("\t row could not save")
                                        #raise Exception(e)

                        #raise Exception(e)
        #------------------------
        return True
    except Exception as e:
        print (f"{filename_to_save} not saved correctly:\t",e)
        return False

#/////////////////////////////////////////////////////
#           Generic File Operations
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
    
def cleanup(filename, input, processed):
    import shutil

    source      = f"{input}{filename}"
    destination = f"{processed}{filename}"

    shutil.move(source,destination)



if __name__=="__main__":
    a = excel_to_list("C:\\Users\\Andrew\\source\\repos\\VENDOR_FILES\\INPUT\\Lotus-322-14.xlsx")
    print(a)