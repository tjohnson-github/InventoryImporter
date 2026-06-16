import dearpygui.dearpygui as dpg
import openpyxl
import os
import gspread
from Google_Sheets_and_Drive_Auth import auth_gspread

def display_quality_greenhouses_transformer(sender,app_data,user_data):

    filepath=user_data

    with dpg.window(label="Transform Quality Greenhouses Spreadsheet",width=600,height=250) as transformer_window:
        
        dpg.add_progress_bar(id=f'{transformer_window}_progress', overlay="Loading products...", show=True, width=200, parent=transformer_window)
        products = parse()
        dpg.configure_item(f'{transformer_window}_progress', show=True, default_value=1, overlay=f"{len(products)} products loaded.")
        
        input_group = dpg.add_group(horizontal=True)
        dpg.add_button(tag=f'{transformer_window}_input_fileselect', label="Choose Input File", callback=file_select, user_data=transformer_window, parent=input_group, width=150)
        dpg.add_input_text(tag=f'{transformer_window}_input_filename', label="Input File", parent=input_group, width=350)
        dpg.add_input_text(tag=f'{transformer_window}_output_filename', label="Output File", width=500)
        dpg.add_button(label="Convert", callback=begin_convert, user_data=products, width=400)
        
def file_select(sender, app_data, user_data):
    with dpg.file_dialog(modal=True, default_path=os.path.join(os.path.expanduser('~'), 'downloads'), callback=update_file_select, height=400, width=400) as dialog:
        dpg.add_file_extension(".xlsx", color=(255, 255, 0, 255), custom_text="XLSX")
        dpg.set_item_user_data(dialog, user_data)

def update_file_select(sender, app_data, user_data):
    sender_parent = user_data
    dpg.set_value(f'{sender_parent}_input_filename', app_data['file_path_name'])
    if not dpg.get_value(f'{sender_parent}_output_filename'):
        output_filename = os.path.join(dpg.get_value('base_input_path'), app_data['file_name'])
        if app_data['file_path_name'] != output_filename:
            dpg.set_value(f'{sender_parent}_output_filename', output_filename)

def parse():

    gc = auth_gspread()
    sh = gc.open("Quality Greenhouses Products")

    locations = None
    products = {}

    for wk in sh.worksheets():
        for row_number, row in enumerate(wk.get_all_values()):
            if locations is None:
                locations = Locations(row)
                print(f"{wk.title}: {locations}")
            else:

                customer_item = normalize(row[locations.customer_item])
                if customer_item is None:
                    raise ValueError(f"{wk.title}: No Customer Item found on row {row_number + 1}")

                retail = normalize(row[locations.retail])
                if retail is None:
                    raise ValueError(f"{wk.title}: No Retail found on row {row_number + 1}")

                products[customer_item] = retail
        
        if len(products) == 0:
            raise ValueError(f"{wk.title}: No products found!")

    print(products)
    return products

def normalize(value):
    if value is None:
        return None
    else:
        stripped = str(value).strip()
        return stripped or None

class Locations:
    customer_item: int
    retail: int

    def __init__(self, row):

        for column_number, cell in enumerate(row):
            name = normalize(cell)
            if name is not None:
                if name == "Customer Item":
                    self.customer_item = column_number
                elif name == "Retail":
                    self.retail = column_number

        if self.customer_item is None:
            raise ValueError("Customer Item column not found!")
        if self.retail is None:
            raise ValueError("Retail column not found!")

    def __repr__(self):
        return f"{self.customer_item} {self.retail}"

def begin_convert(sender, app_data, user_data):

    transformer_window = dpg.get_item_parent(sender)
    products = user_data

    input_filename = dpg.get_value(f'{transformer_window}_input_filename')
    output_filename = dpg.get_value(f'{transformer_window}_output_filename')
    progress_bar_id = f'{transformer_window}_progress'

    dpg.configure_item(progress_bar_id, show=True, default_value=0, overlay=f"0%")

    wb = openpyxl.load_workbook(input_filename, data_only=True)
    ws = wb.active  

    output_wb = openpyxl.Workbook()
    output = output_wb.active

    dpg.configure_item(progress_bar_id,show=True,default_value=0.1,overlay="10%")

    first_row = True
    for row in ws.iter_rows():
        if first_row:
            output.append([
                "Description",
                "Man #",
                "Qty",
                "Cost",
                "Retail",
                "UPC"
            ])
            first_row = False
        else:
            upc = normalize(row[6].value).replace('-', '')
            output.append([
                normalize(row[0].value),
                normalize(row[1].value),
                normalize(row[3].value),
                normalize(row[5].value),
                products.get(upc, ''),
                upc
            ])

    dpg.configure_item(progress_bar_id,show=True,default_value=0.9,overlay="90%")
    output_wb.save(output_filename)
    dpg.configure_item(progress_bar_id,show=True,default_value=1,overlay="Done")