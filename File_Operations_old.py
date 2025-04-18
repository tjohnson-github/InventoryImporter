
import pickle
import openpyxl
import csv
import sys,os
import pandas as pd
import tabula


#/////////////////////////////////////////////////////
#           PICKLING
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\

def saveVariable(save_Name,save_Data):
    #--------------
    try:
        pickle.dump(save_Data,open(save_Name,'wb'))
    except Exception as e:
        print (f"{saveName} not saved correctly. Check permissions")
        print (e)
        return e
    
def getVariable(save_Name):
    #--------------
    try:
        saved_Data = pickle.load(open(save_Name,'rb'))
        return saved_Data
    except Exception as e:
        print (f"{saveName} not loaded correctly. Check permissions")
        print (e)
        return e
    
#/////////////////////////////////////////////////////
#           EXCEL / CSV <--> LIST CONVERSIONS
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
    
def csv_to_list(csv_file,delimiter = ','):
	#-------------------------------------------
    if os.path.exists(csv_file)==False:     
        error = f"File Path does not exist for item: {str(csv_file)}."
        return False,error
    elif csv_file.endswith(".csv")==False:
        error = f"{str(csv_file)} is not correct file type for this operation; .csv only please."
        return False,error
    #else:
    #    error = f"Unknown error while attempting to read {csv_file}."
    #    return False,error
	#-------------------------------------------
    readArray=[]
    with open(csv_file,newline='') as csvfile:
        contents=csv.reader(csvfile, delimiter=delimiter, quotechar='"')
        for row in contents:
            temp_array=[]
            #----------
            for item in row:
                #============ TRIM WHITESPACES!
                temp_value = str(item)
                while temp_value.endswith(' '):
                    temp_value = temp_value[:-1]
                temp_array.append(item)
                #============ TRIM WHITESPACES!
            #----------
            if (temp_array != [None for i in range(0,len(temp_array))]) and (temp_array != ['' for i in range(0,len(temp_array))]):
                readArray.append(temp_array)
	#-------------------------------------------
    return readArray,''


def excel_to_list(excel_file):
	#-------------------------------------------
    if os.path.exists(excel_file)==False:   
        error = f"File Path does not exist for item: {str(excel_file)}."
        print (error)
        return False,error
    elif not excel_file.endswith('.xlsx'):
        error = f"{str(excel_file)} is not correct file type for this operation; .xlsx only please."
        print (error)
        return False,error
    #else:
    #    error = f"Unknown error while attempting to read {excel_file}."
    #    return False,error
	#-------------------------------------------
    wb = openpyxl.load_workbook(excel_file,data_only=True)
    ws = wb.active
	#-------------------------------------------
    full_list=[]
	#-------------------------------------------
    for row in ws.iter_rows():
        #----------------
        temp_list=[]
        #----------------
        for cell in row:
            #----------------
            if cell.value==' ':     
                temp_list.append('')
            else:                   
                #============ TRIM WHITESPACES!
                temp_value = str(cell.value)
                while temp_value.endswith(' '):
                    temp_value = temp_value[:-1]
                temp_list.append(temp_value)
                #============ TRIM WHITESPACES!

        #----------------
        if (temp_list != [None for i in range(0,len(temp_list))]) and (temp_list != ['' for i in range(0,len(temp_list))]):
            full_list.append(temp_list)
	#-------------------------------------------
    return full_list,''


def list_to_excel(temp_list,filename_to_save):
    import openpyxl
    try:
        #------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        #------------------------
        for row in temp_list:
            ws.append(row)
        #------------------------
        wb.save(filename_to_save)
        return True
    except Exception as e:
        print (f"{filename_to_save} not saved correctly:\t{e}")
        return False

def list_to_csv(temp_list, filename_to_save, delimiter = ','):
    try:
        #------------------------
        with open(filename_to_save, 'w', newline='') as csvfile:
            writer_object = csv.writer(csvfile, delimiter = delimiter, quotechar='"', quoting = csv.QUOTE_MINIMAL)
            #------------------------
            for row in temp_list:
                writer_object.writerow(row)
        #------------------------
        return True
    except Exception as e:
        print (f"{filename_to_save} not saved correctly:\t",e)
        return False

#/////////////////////////////////////////////////////
#           Generic File Operations
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
    
def cleanup(filename, input, processed):
    import shutil

    source      = f"{input}\\{filename}"
    destination = f"{processed}\\{filename}"

    shutil.move(source,destination)

